"""
پنجره اصلی رابط کاربری - راست‌چین
"""

import sys
import os
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from pathlib import Path

from ui.styles import MAIN_STYLE, LOG_STYLE
from src.converter import XMLConverter


class MainWindow(QMainWindow):
    """
    پنجره اصلی برنامه - راست‌چین
    """
    
    def __init__(self):
        super().__init__()
        self.converter = XMLConverter()
        self.current_file = None
        self.init_ui()
        
    def init_ui(self):
        """
        راه‌اندازی رابط کاربری
        """
        self.setWindowTitle("XML to Excel Converter")
        self.setWindowIcon(self.create_icon())
        self.setGeometry(100, 100, 900, 700)
        self.setStyleSheet(MAIN_STYLE)
        
        # تنظیم جهت راست‌چین برای کل برنامه
        self.setLayoutDirection(Qt.RightToLeft)
        
        # ویجت مرکزی
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # لایه‌بندی اصلی
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)
        
        # فریم اصلی
        self.main_frame = QFrame()
        self.main_frame.setObjectName("main_frame")
        main_layout.addWidget(self.main_frame)
        
        # لایه‌بندی فریم
        frame_layout = QVBoxLayout(self.main_frame)
        frame_layout.setSpacing(15)
        
        # هدر
        self.create_header(frame_layout)
        
        # بخش انتخاب فایل
        self.create_file_section(frame_layout)
        
        # بخش تنظیمات
        self.create_settings_section(frame_layout)
        
        # بخش خروجی
        self.create_output_section(frame_layout)
        
        # بخش لاگ
        self.create_log_section(frame_layout)
        
        # بخش دکمه‌ها
        self.create_buttons_section(frame_layout)
        
        # تنظیم نسبت‌ها
        frame_layout.setStretchFactor(self.log_text, 2)
        
    def create_icon(self):
        """
        ایجاد آیکون برنامه
        """
        pixmap = QPixmap(64, 64)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # رسم آیکون ساده
        painter.setBrush(QColor(102, 126, 234))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(10, 10, 44, 44, 10, 10)
        
        painter.setPen(QPen(Qt.white, 3))
        painter.setBrush(Qt.NoBrush)
        painter.drawRoundedRect(16, 16, 32, 32, 5, 5)
        
        painter.setPen(QPen(Qt.white, 2))
        painter.drawLine(24, 28, 40, 28)
        painter.drawLine(24, 32, 40, 32)
        painter.drawLine(24, 36, 35, 36)
        
        painter.end()
        return QIcon(pixmap)
    
    def create_header(self, layout):
        """
        ایجاد بخش هدر - راست‌چین
        """
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setDirection(QHBoxLayout.RightToLeft)  # راست‌چین
        
        title_label = QLabel("📊 XML to Excel Converter")
        title_label.setObjectName("title_label")
        title_label.setAlignment(Qt.AlignRight)  # راست‌چین
        header_layout.addWidget(title_label)
        
        version_label = QLabel("v2.0")
        version_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        header_layout.addWidget(version_label)
        header_layout.addStretch()
        
        layout.addWidget(header_widget)
        
        subtitle = QLabel("تبدیل فایل‌های XML به Excel، CSV و JSON با رابط کاربری زیبا")
        subtitle.setObjectName("subtitle_label")
        subtitle.setAlignment(Qt.AlignRight)  # راست‌چین
        layout.addWidget(subtitle)
    
    def create_file_section(self, layout):
        """
        ایجاد بخش انتخاب فایل - راست‌چین
        """
        file_group = QGroupBox("📁 فایل ورودی")
        file_group.setLayoutDirection(Qt.RightToLeft)
        file_layout = QVBoxLayout()
        file_layout.setDirection(QVBoxLayout.TopToBottom)
        
        # انتخاب فایل
        file_select_layout = QHBoxLayout()
        file_select_layout.setDirection(QHBoxLayout.RightToLeft)  # راست‌چین
        
        self.file_path = QLineEdit()
        self.file_path.setPlaceholderText("مسیر فایل XML را انتخاب کنید...")
        self.file_path.setReadOnly(True)
        self.file_path.setAlignment(Qt.AlignRight)  # راست‌چین
        file_select_layout.addWidget(self.file_path)
        
        browse_btn = QPushButton("📂 انتخاب فایل")
        browse_btn.clicked.connect(self.browse_file)
        file_select_layout.addWidget(browse_btn)
        
        file_layout.addLayout(file_select_layout)
        
        # اطلاعات فایل
        self.file_info = QLabel("هیچ فایلی انتخاب نشده است")
        self.file_info.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        self.file_info.setAlignment(Qt.AlignRight)  # راست‌چین
        file_layout.addWidget(self.file_info)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
    
    def create_settings_section(self, layout):
        """
        ایجاد بخش تنظیمات - راست‌چین
        """
        settings_group = QGroupBox("⚙️ تنظیمات")
        settings_group.setLayoutDirection(Qt.RightToLeft)
        settings_layout = QGridLayout()
        settings_layout.setSpacing(15)
        
        # تگ هدف
        label_tag = QLabel("🎯 تگ هدف:")
        label_tag.setAlignment(Qt.AlignRight)  # راست‌چین
        settings_layout.addWidget(label_tag, 0, 0)
        
        self.tag_input = QLineEdit()
        self.tag_input.setPlaceholderText("مثال: food (اختیاری)")
        self.tag_input.setAlignment(Qt.AlignRight)  # راست‌چین
        settings_layout.addWidget(self.tag_input, 0, 1)
        
        # فرمت خروجی
        label_format = QLabel("📄 فرمت خروجی:")
        label_format.setAlignment(Qt.AlignRight)  # راست‌چین
        settings_layout.addWidget(label_format, 1, 0)
        
        self.format_combo = QComboBox()
        self.format_combo.addItems(["Excel (xlsx)", "CSV", "JSON", "همه فرمت‌ها"])
        settings_layout.addWidget(self.format_combo, 1, 1)
        
        # نام فایل خروجی
        label_name = QLabel("📝 نام فایل:")
        label_name.setAlignment(Qt.AlignRight)  # راست‌چین
        settings_layout.addWidget(label_name, 2, 0)
        
        self.output_name = QLineEdit()
        self.output_name.setPlaceholderText("نام فایل خروجی (بدون پسوند)")
        self.output_name.setText("converted")
        self.output_name.setAlignment(Qt.AlignRight)  # راست‌چین
        settings_layout.addWidget(self.output_name, 2, 1)
        
        # مسیر خروجی
        label_path = QLabel("📁 مسیر خروجی:")
        label_path.setAlignment(Qt.AlignRight)  # راست‌چین
        settings_layout.addWidget(label_path, 3, 0)
        
        output_path_layout = QHBoxLayout()
        output_path_layout.setDirection(QHBoxLayout.RightToLeft)  # راست‌چین
        
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("مسیر ذخیره فایل...")
        self.output_path.setText(os.path.join(os.getcwd(), "output"))
        self.output_path.setAlignment(Qt.AlignRight)  # راست‌چین
        output_path_layout.addWidget(self.output_path)
        
        browse_output_btn = QPushButton("📂 انتخاب")
        browse_output_btn.clicked.connect(self.browse_output)
        browse_output_btn.setFixedWidth(80)
        output_path_layout.addWidget(browse_output_btn)
        
        settings_layout.addLayout(output_path_layout, 3, 1)
        
        # گزینه‌های اضافی
        options_layout = QHBoxLayout()
        options_layout.setDirection(QHBoxLayout.RightToLeft)  # راست‌چین
        
        self.summary_check = QCheckBox("📊 تولید گزارش خلاصه")
        self.summary_check.setChecked(True)
        options_layout.addWidget(self.summary_check)
        
        self.preview_check = QCheckBox("👁️ نمایش پیش‌نمایش")
        self.preview_check.setChecked(True)
        options_layout.addWidget(self.preview_check)
        
        settings_layout.addLayout(options_layout, 4, 0, 1, 2)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
    
    def create_output_section(self, layout):
        """
        ایجاد بخش خروجی - راست‌چین
        """
        output_group = QGroupBox("📊 پیش‌نمایش داده")
        output_group.setLayoutDirection(Qt.RightToLeft)
        output_layout = QVBoxLayout()
        
        self.preview_text = QTextEdit()
        self.preview_text.setPlaceholderText("پس از تبدیل، داده‌ها در اینجا نمایش داده می‌شوند...")
        self.preview_text.setMaximumHeight(150)
        self.preview_text.setAlignment(Qt.AlignRight)  # راست‌چین
        output_layout.addWidget(self.preview_text)
        
        output_group.setLayout(output_layout)
        layout.addWidget(output_group)
    
    def create_log_section(self, layout):
        """
        ایجاد بخش لاگ - راست‌چین
        """
        log_group = QGroupBox("📋 گزارش عملیات")
        log_group.setLayoutDirection(Qt.RightToLeft)
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setPlaceholderText("گزارش عملیات در اینجا نمایش داده می‌شود...")
        self.log_text.setStyleSheet(LOG_STYLE)
        self.log_text.setReadOnly(True)
        self.log_text.setAlignment(Qt.AlignRight)  # راست‌چین
        log_layout.addWidget(self.log_text)
        
        log_group.setLayout(log_layout)
        layout.addWidget(log_group)
    
    def create_buttons_section(self, layout):
        """
        ایجاد بخش دکمه‌ها - راست‌چین
        """
        button_layout = QHBoxLayout()
        button_layout.setDirection(QHBoxLayout.RightToLeft)  # راست‌چین
        button_layout.setSpacing(10)
        
        # دکمه تبدیل
        self.convert_btn = QPushButton("🔄 شروع تبدیل")
        self.convert_btn.setObjectName("success_btn")
        self.convert_btn.setFixedHeight(45)
        self.convert_btn.clicked.connect(self.convert_file)
        self.convert_btn.setEnabled(False)
        button_layout.addWidget(self.convert_btn)
        
        # دکمه پاک کردن
        clear_btn = QPushButton("🗑️ پاک کردن لاگ")
        clear_btn.setObjectName("danger_btn")
        clear_btn.setFixedHeight(45)
        clear_btn.clicked.connect(self.clear_log)
        button_layout.addWidget(clear_btn)
        
        # دکمه خروج
        exit_btn = QPushButton("🚪 خروج")
        exit_btn.setFixedHeight(45)
        exit_btn.clicked.connect(self.close)
        button_layout.addWidget(exit_btn)
        
        layout.addLayout(button_layout)
    
    def browse_file(self):
        """
        انتخاب فایل XML
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "انتخاب فایل XML",
            "",
            "فایل‌های XML (*.xml);;همه فایل‌ها (*.*)"
        )
        
        if file_path:
            self.current_file = file_path
            self.file_path.setText(file_path)
            self.convert_btn.setEnabled(True)
            
            # نمایش اطلاعات فایل
            file_size = os.path.getsize(file_path)
            size_str = self.format_size(file_size)
            self.file_info.setText(f"📄 {os.path.basename(file_path)} - حجم: {size_str}")
            
            # بارگذاری و نمایش ساختار
            self.load_xml_info(file_path)
            
            self.log_message(f"فایل انتخاب شد: {file_path}", "INFO")
    
    def browse_output(self):
        """
        انتخاب مسیر خروجی
        """
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "انتخاب مسیر خروجی",
            self.output_path.text()
        )
        
        if dir_path:
            self.output_path.setText(dir_path)
            self.log_message(f"مسیر خروجی انتخاب شد: {dir_path}", "INFO")
    
    def load_xml_info(self, file_path):
        """
        بارگذاری اطلاعات XML
        """
        try:
            info = self.converter.get_xml_info(file_path)
            
            preview_text = f"""
📊 اطلاعات ساختار XML:

🏷️ تگ ریشه: {info['root_tag']}
📦 تعداد کل المان‌ها: {info['total_elements']}
📏 حداکثر عمق: {info['max_depth']}
🔖 تگ‌های منحصر‌به‌فرد: {', '.join(info['unique_tags'][:10])}
            """
            
            if len(info['unique_tags']) > 10:
                preview_text += f"\n   و {len(info['unique_tags']) - 10} تگ دیگر..."
            
            self.preview_text.setText(preview_text)
            
        except Exception as e:
            self.log_message(f"خطا در خواندن اطلاعات XML: {e}", "ERROR")
    
    def convert_file(self):
        """
        تبدیل فایل XML
        """
        if not self.current_file:
            QMessageBox.warning(self, "هشدار", "لطفاً ابتدا یک فایل XML انتخاب کنید!")
            return
        
        try:
            # غیرفعال کردن دکمه
            self.convert_btn.setEnabled(False)
            self.convert_btn.setText("⏳ در حال تبدیل...")
            
            # دریافت تنظیمات
            tag = self.tag_input.text().strip() if self.tag_input.text().strip() else None
            output_format = self.format_combo.currentText()
            output_name = self.output_name.text().strip() or "converted"
            output_dir = self.output_path.text()
            generate_summary = self.summary_check.isChecked()
            
            # ایجاد دایرکتوری خروجی
            os.makedirs(output_dir, exist_ok=True)
            
            self.log_message(f"شروع تبدیل فایل: {os.path.basename(self.current_file)}", "INFO")
            self.log_message(f"فرمت خروجی: {output_format}", "INFO")
            
            # تبدیل
            results = self.converter.convert(
                xml_file=self.current_file,
                output_dir=output_dir,
                output_name=output_name,
                target_tag=tag,
                output_format=output_format,
                generate_summary=generate_summary
            )
            
            # نمایش نتایج
            self.log_message("✅ تبدیل با موفقیت انجام شد!", "SUCCESS")
            
            for format_name, file_path in results.items():
                self.log_message(f"📁 {format_name}: {file_path}", "SUCCESS")
            
            # نمایش پیش‌نمایش
            if self.preview_check.isChecked() and results:
                self.show_preview(results)
            
            # پیام موفقیت
            QMessageBox.information(
                self,
                "موفقیت",
                f"✅ فایل با موفقیت تبدیل شد!\n\n"
                f"تعداد فایل‌های ایجاد شده: {len(results)}\n"
                f"مسیر: {output_dir}"
            )
            
        except Exception as e:
            self.log_message(f"❌ خطا در تبدیل: {e}", "ERROR")
            QMessageBox.critical(self, "خطا", f"خطا در تبدیل فایل:\n{str(e)}")
        
        finally:
            self.convert_btn.setEnabled(True)
            self.convert_btn.setText("🔄 شروع تبدیل")
    
    def show_preview(self, results):
        """
        نمایش پیش‌نمایش داده‌ها (بدون Pandas)
        """
        try:
            from openpyxl import load_workbook
            
            # پیدا کردن فایل Excel
            excel_file = None
            for format_name, file_path in results.items():
                if format_name == "Excel (xlsx)":
                    excel_file = file_path
                    break
            
            if excel_file and os.path.exists(excel_file):
                # خواندن فایل Excel با OpenPyXL
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                
                # استخراج داده‌ها
                headers = []
                data = []
                
                # خواندن هدرها (سطر اول)
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                
                # خواندن داده‌ها (حداکثر 10 سطر)
                max_rows = min(ws.max_row, 11)
                for row in range(2, max_rows + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row, column=col_idx).value
                        row_data[header] = cell_value if cell_value is not None else ''
                    data.append(row_data)
                
                # ساخت متن پیش‌نمایش
                preview_text = f"""
📊 پیش‌نمایش داده‌ها:

تعداد کل رکوردها: {ws.max_row - 1}
تعداد ستون‌ها: {len(headers)}
ستون‌ها: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}

نمونه داده‌ها (حداکثر 10 رکورد):
"""
                # اضافه کردن داده‌ها به صورت جدول
                if data:
                    preview_text += "\n" + " | ".join(headers[:8]) + "\n"
                    preview_text += "-" * 50 + "\n"
                    
                    for row_data in data[:10]:
                        row_values = []
                        for header in headers[:8]:
                            value = row_data.get(header, '')
                            if value is None:
                                value = ''
                            str_value = str(value)[:20]
                            if len(str(value)) > 20:
                                str_value += '...'
                            row_values.append(str_value)
                        preview_text += " | ".join(row_values) + "\n"
                
                self.preview_text.setText(preview_text)
                
        except Exception as e:
            self.log_message(f"خطا در نمایش پیش‌نمایش: {e}", "WARNING")
            self.preview_text.setText("⚠️ پیش‌نمایش در دسترس نیست (برای مشاهده فایل Excel را باز کنید)")
    
    def log_message(self, message, level="INFO"):
        """
        افزودن پیام به لاگ - راست‌چین
        """
        timestamp = QDateTime.currentDateTime().toString("hh:mm:ss")
        
        # تنظیم رنگ بر اساس سطح
        if level == "ERROR":
            formatted = f'<span style="color: #e57373;">[{timestamp}] ❌ {message}</span>'
        elif level == "WARNING":
            formatted = f'<span style="color: #ffb74d;">[{timestamp}] ⚠️ {message}</span>'
        elif level == "SUCCESS":
            formatted = f'<span style="color: #81c784;">[{timestamp}] ✅ {message}</span>'
        else:
            formatted = f'<span style="color: #4fc3f7;">[{timestamp}] ℹ️ {message}</span>'
        
        self.log_text.append(formatted)
        # اسکرول به پایین
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
    
    def clear_log(self):
        """
        پاک کردن لاگ
        """
        self.log_text.clear()
        self.log_message("لاگ پاک شد", "INFO")
    
    def format_size(self, size):
        """
        فرمت کردن حجم فایل
        """
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024.0:
                return f"{size:.2f} {unit}"
            size /= 1024.0
        return f"{size:.2f} TB"


def run_app():
    """
    اجرای برنامه
    """
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    # تنظیم جهت برنامه به راست‌چین
    app.setLayoutDirection(Qt.RightToLeft)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    run_app()