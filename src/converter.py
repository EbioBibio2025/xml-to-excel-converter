"""
ماژول اصلی تبدیل XML به فرمت‌های مختلف (نسخه کاملاً مستقل)
"""

import os
import json
import csv
from typing import Dict, List, Any, Optional
import logging
from xml.etree import ElementTree as ET

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class XMLConverter:
    """
    کلاس اصلی تبدیل XML به فرمت‌های مختلف - کاملاً مستقل از Pandas
    """
    
    def __init__(self):
        self.data = []
        self.headers = []
        self.parser = None
    
    def get_xml_info(self, xml_file: str) -> Dict:
        """
        دریافت اطلاعات ساختاری XML
        """
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            # پیدا کردن تگ‌های منحصر به فرد
            unique_tags = set()
            max_depth = 0
            
            def traverse(element, depth=0):
                nonlocal max_depth
                unique_tags.add(element.tag)
                max_depth = max(max_depth, depth)
                for child in element:
                    traverse(child, depth + 1)
            
            traverse(root)
            
            return {
                'root_tag': root.tag,
                'total_elements': len(list(root.iter())),
                'unique_tags': list(unique_tags),
                'max_depth': max_depth
            }
        except Exception as e:
            logger.error(f"خطا در خواندن اطلاعات XML: {e}")
            raise
    
    def _extract_data(self, xml_file: str, target_tag: Optional[str] = None) -> List[Dict]:
        """
        استخراج داده‌ها از XML
        """
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        data = []
        
        # پیدا کردن المان‌های هدف
        if target_tag:
            elements = root.findall(f'.//{target_tag}')
        else:
            elements = [root]
        
        for element in elements:
            row = {}
            
            # اضافه کردن ویژگی‌ها
            for attr_name, attr_value in element.attrib.items():
                row[f'@{attr_name}'] = attr_value
            
            # اضافه کردن زیرالمان‌ها
            for child in element:
                if len(child) == 0:  # برگ
                    if child.tag in row:
                        # اگر چندین مقدار با تگ یکسان داریم
                        if not isinstance(row[child.tag], list):
                            row[child.tag] = [row[child.tag]]
                        row[child.tag].append(child.text)
                    else:
                        row[child.tag] = child.text
                else:  # دارای زیرالمان
                    if child.tag not in row:
                        row[child.tag] = {}
                    for grandchild in child:
                        row[child.tag][grandchild.tag] = grandchild.text
            
            data.append(row)
        
        return data
    
    def _flatten_data(self, data: List[Dict]) -> List[Dict]:
        """
        مسطح کردن داده‌های تو در تو
        """
        flat_data = []
        
        for item in data:
            flat_row = {}
            
            def flatten(d, parent_key=''):
                for k, v in d.items():
                    new_key = f"{parent_key}_{k}" if parent_key else k
                    
                    if isinstance(v, dict):
                        flatten(v, new_key)
                    elif isinstance(v, list):
                        # برای لیست‌ها، آنها را به رشته تبدیل کن
                        flat_row[new_key] = ', '.join(str(item) for item in v)
                    else:
                        flat_row[new_key] = v
            
            flatten(item)
            flat_data.append(flat_row)
        
        return flat_data
    
    def convert(self, xml_file: str, output_dir: str = "output", 
                output_name: str = "converted", target_tag: Optional[str] = None,
                output_format: str = "Excel (xlsx)", generate_summary: bool = True) -> Dict[str, str]:
        """
        تبدیل فایل XML به فرمت‌های مختلف
        """
        results = {}
        
        # ایجاد دایرکتوری خروجی
        os.makedirs(output_dir, exist_ok=True)
        
        # استخراج داده‌ها
        raw_data = self._extract_data(xml_file, target_tag)
        self.data = self._flatten_data(raw_data)
        
        if not self.data:
            raise ValueError("هیچ داده‌ای برای تبدیل یافت نشد!")
        
        # استخراج هدرها
        if self.data:
            self.headers = list(self.data[0].keys())
        
        logger.info(f"{len(self.data)} رکورد برای تبدیل آماده شد")
        
        # تبدیل به فرمت‌های مختلف
        if output_format in ["Excel (xlsx)", "همه فرمت‌ها"]:
            excel_file = self._convert_to_excel(output_dir, output_name)
            results["Excel (xlsx)"] = excel_file
        
        if output_format in ["CSV", "همه فرمت‌ها"]:
            csv_file = self._convert_to_csv(output_dir, output_name)
            results["CSV"] = csv_file
        
        if output_format in ["JSON", "همه فرمت‌ها"]:
            json_file = self._convert_to_json(output_dir, output_name)
            results["JSON"] = json_file
        
        # گزارش خلاصه
        if generate_summary:
            summary_file = self._generate_summary(output_dir, output_name)
            results["گزارش خلاصه"] = summary_file
        
        return results
    
    def _convert_to_excel(self, output_dir: str, output_name: str) -> str:
        """
        تبدیل به Excel با OpenPyXL
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        filepath = os.path.join(output_dir, f"{output_name}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "داده‌های XML"
        
        # نوشتن هدرها
        for col_idx, header in enumerate(self.headers, 1):
            ws.cell(row=1, column=col_idx, value=str(header))
        
        # نوشتن داده‌ها
        for row_idx, row_data in enumerate(self.data, 2):
            for col_idx, header in enumerate(self.headers, 1):
                value = row_data.get(header, '')
                if isinstance(value, (dict, list)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # اعمال استایل
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # استایل هدرها
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # استایل سلول‌ها
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # تنظیم عرض ستون‌ها
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        logger.info(f"فایل Excel ایجاد شد: {filepath}")
        return filepath
    
    def _convert_to_csv(self, output_dir: str, output_name: str) -> str:
        """
        تبدیل به CSV
        """
        filepath = os.path.join(output_dir, f"{output_name}.csv")
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(self.headers)
            
            for row_data in self.data:
                row = [str(row_data.get(header, '')) for header in self.headers]
                writer.writerow(row)
        
        logger.info(f"فایل CSV ایجاد شد: {filepath}")
        return filepath
    
    def _convert_to_json(self, output_dir: str, output_name: str) -> str:
        """
        تبدیل به JSON
        """
        filepath = os.path.join(output_dir, f"{output_name}.json")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"فایل JSON ایجاد شد: {filepath}")
        return filepath
    
    def _generate_summary(self, output_dir: str, output_name: str) -> str:
        """
        تولید گزارش خلاصه
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        filepath = os.path.join(output_dir, f"{output_name}_summary.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "خلاصه"
        
        summary_data = [
            ['معیار', 'مقدار'],
            ['تعداد کل رکوردها', len(self.data)],
            ['تعداد ستون‌ها', len(self.headers)],
            ['نام ستون‌ها', ', '.join(self.headers[:10]) + ('...' if len(self.headers) > 10 else '')],
            ['تاریخ ایجاد', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        logger.info(f"گزارش خلاصه ایجاد شد: {filepath}")
        return filepath