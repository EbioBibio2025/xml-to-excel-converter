"""
ماژول تولید فایل‌های Excel
این ماژول مسئول تبدیل داده‌های استخراج شده به فایل Excel است.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Optional
import logging
from datetime import datetime
import os

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    کلاس تولید فایل‌های Excel با قابلیت‌های مختلف
    """
    
    def __init__(self, output_dir: str = 'data/output'):
        """
        سازنده کلاس
        
        Args:
            output_dir (str): دایرکتوری خروجی
        """
        self.output_dir = output_dir
        self._create_output_directory()
    
    def _create_output_directory(self) -> None:
        """
        ایجاد دایرکتوری خروجی اگر وجود نداشته باشد
        """
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"دایرکتوری خروجی ایجاد شد: {self.output_dir}")
    
    def generate_with_pandas(self, data: List[Dict], filename: str = 'output.xlsx') -> str:
        """
        تولید فایل Excel با استفاده از Pandas
        
        Args:
            data (List[Dict]): لیست دیکشنری‌های داده
            filename (str): نام فایل خروجی
            
        Returns:
            str: مسیر فایل ایجاد شده
        """
        try:
            # تبدیل به DataFrame
            df = pd.DataFrame(data)
            
            # پر کردن مقادیر None با رشته خالی
            df = df.fillna('')
            
            # مسیر کامل فایل
            filepath = os.path.join(self.output_dir, filename)
            
            # ذخیره در Excel با تنظیمات
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='داده‌ها', index=False)
                
                # دریافت workbook برای اعمال استایل
                workbook = writer.book
                worksheet = writer.sheets['داده‌ها']
                
                # اعمال استایل به هدرها
                self._style_header(worksheet)
                
                # تنظیم عرض ستون‌ها
                self._auto_fit_columns(worksheet)
            
            logger.info(f"فایل Excel با Pandas ایجاد شد: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"خطا در تولید Excel با Pandas: {e}")
            raise
    
    def generate_with_openpyxl(self, data: List[Dict], filename: str = 'styled_output.xlsx') -> str:
        """
        تولید فایل Excel با استایل‌دهی پیشرفته با OpenPyXL
        
        Args:
            data (List[Dict]): لیست دیکشنری‌های داده
            filename (str): نام فایل خروجی
            
        Returns:
            str: مسیر فایل ایجاد شده
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "داده‌های XML"
            
            # اگر داده وجود ندارد
            if not data:
                ws['A1'] = "هیچ داده‌ای برای نمایش وجود ندارد"
                filepath = os.path.join(self.output_dir, filename)
                wb.save(filepath)
                return filepath
            
            # نوشتن هدرها
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
            
            # نوشتن داده‌ها
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # اعمال استایل‌ها
            self._apply_styles(ws, len(data), len(headers))
            
            # ذخیره فایل
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            logger.info(f"فایل Excel با OpenPyXL و استایل ایجاد شد: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"خطا در تولید Excel با OpenPyXL: {e}")
            raise
    
    def _apply_styles(self, worksheet, row_count: int, col_count: int) -> None:
        """
        اعمال استایل‌های مختلف به شیت
        
        Args:
            worksheet: شیت OpenPyXL
            row_count (int): تعداد سطرها
            col_count (int): تعداد ستون‌ها
        """
        # تعریف استایل‌ها
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # استایل هدرها
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # استایل سلول‌های داده
        for row in range(2, row_count + 2):
            for col in range(1, col_count + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # تنظیم عرض ستون‌ها
        self._auto_fit_columns(worksheet)
    
    def _style_header(self, worksheet) -> None:
        """
        اعمال استایل به هدرهای شیت
        
        Args:
            worksheet: شیت OpenPyXL
        """
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    def _auto_fit_columns(self, worksheet) -> None:
        """
        تنظیم خودکار عرض ستون‌ها بر اساس محتوا
        
        Args:
            worksheet: شیت OpenPyXL
        """
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    def generate_summary_report(self, data: List[Dict], filename: str = 'summary.xlsx') -> str:
        """
        تولید گزارش خلاصه از داده‌ها
        
        Args:
            data (List[Dict]): لیست دیکشنری‌های داده
            filename (str): نام فایل خروجی
            
        Returns:
            str: مسیر فایل ایجاد شده
        """
        try:
            df = pd.DataFrame(data)
            
            # ایجاد گزارش خلاصه
            summary_data = {
                'معیار': [
                    'تعداد کل رکوردها',
                    'تعداد ستون‌ها',
                    'نام ستون‌ها',
                    'تاریخ ایجاد',
                    'تعداد مقادیر خالی'
                ],
                'مقدار': [
                    len(df),
                    len(df.columns),
                    ', '.join(df.columns.tolist()),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    df.isnull().sum().sum()
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            
            filepath = os.path.join(self.output_dir, filename)
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='خلاصه', index=False)
                df.to_excel(writer, sheet_name='داده‌های کامل', index=False)
            
            logger.info(f"گزارش خلاصه ایجاد شد: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"خطا در تولید گزارش خلاصه: {e}")
            raise


if __name__ == "__main__":
    # تست ماژول
    test_data = [
        {'نام': 'پیتزا', 'قیمت': 150000, 'کالری': 850},
        {'نام': 'سالاد', 'قیمت': 120000, 'کالری': 420}
    ]
    
    generator = ExcelGenerator()
    generator.generate_with_pandas(test_data, 'test_output.xlsx')
    generator.generate_with_openpyxl(test_data, 'test_styled.xlsx')