"""
ماژول اصلی تبدیل XML به فرمت‌های مختلف - نسخه نهایی با رفع مشکل لیست
"""

import os
import json
import csv
from typing import Dict, List, Any, Optional
import logging
from xml.etree import ElementTree as ET
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class XMLConverter:
    """
    کلاس اصلی تبدیل XML به فرمت‌های مختلف
    """
    
    def __init__(self):
        self.data = []
        self.headers = []
    
    def _is_spreadsheet_xml(self, xml_file: str) -> bool:
        """تشخیص فایل SpreadsheetXML (اکسل)"""
        try:
            with open(xml_file, 'r', encoding='utf-8') as f:
                first_line = f.readline()
                if 'mso-application' in first_line:
                    return True
                content = f.read(500)
                if 'urn:schemas-microsoft-com:office:spreadsheet' in content:
                    return True
                if 'Excel.Sheet' in content:
                    return True
            return False
        except:
            return False
    
    def _extract_from_spreadsheet_xml(self, xml_file: str) -> List[Dict]:
        """استخراج از SpreadsheetXML"""
        try:
            namespaces = {
                'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
                'o': 'urn:schemas-microsoft-com:office:office',
                'x': 'urn:schemas-microsoft-com:office:excel',
            }
            
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            table = root.find('.//ss:Table', namespaces)
            if table is None:
                return []
            
            rows = table.findall('ss:Row', namespaces)
            if not rows:
                return []
            
            # خواندن هدرها
            headers = []
            first_row = rows[0]
            for cell in first_row.findall('ss:Cell', namespaces):
                data_elem = cell.find('ss:Data', namespaces)
                headers.append(data_elem.text.strip() if data_elem is not None else f"col_{len(headers)+1}")
            
            # خواندن داده‌ها
            data = []
            for row in rows[1:]:
                cells = row.findall('ss:Cell', namespaces)
                row_data = {}
                for idx, cell in enumerate(cells):
                    if idx < len(headers):
                        data_elem = cell.find('ss:Data', namespaces)
                        if data_elem is not None and data_elem.text:
                            value = data_elem.text.strip()
                            try:
                                value = int(value) if '.' not in value else float(value)
                            except ValueError:
                                pass
                            row_data[headers[idx]] = value
                if row_data:
                    data.append(row_data)
            
            logger.info(f"{len(data)} رکورد از SpreadsheetXML استخراج شد")
            return data
            
        except Exception as e:
            logger.error(f"خطا در استخراج از SpreadsheetXML: {e}")
            return []
    
    def _extract_normal_xml(self, xml_file: str, target_tag: Optional[str] = None) -> List[Dict]:
        """
        استخراج از XML معمولی - با تشخیص خودکار المان‌های تکراری
        """
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            data = []
            
            # اگر تگ هدف مشخص شده
            if target_tag:
                elements = root.findall(f'.//{target_tag}')
                for element in elements:
                    row = self._parse_element(element)
                    if row:
                        data.append(row)
                logger.info(f"{len(data)} رکورد با تگ '{target_tag}' استخراج شد")
                return data
            
            # پیدا کردن تگ‌های تکراری در سطح اول
            tag_counts = {}
            for child in root:
                tag_counts[child.tag] = tag_counts.get(child.tag, 0) + 1
            
            # پیدا کردن تگ‌هایی که بیش از یک بار تکرار شده‌اند
            repeated_tags = [tag for tag, count in tag_counts.items() if count > 1]
            
            if repeated_tags:
                # استفاده از اولین تگ تکراری
                main_tag = repeated_tags[0]
                elements = root.findall(f'.//{main_tag}')
                logger.info(f"✅ تگ تکراری '{main_tag}' با {len(elements)} تکرار پیدا شد")
                
                for element in elements:
                    row = self._parse_element(element)
                    if row:
                        data.append(row)
                
                logger.info(f"{len(data)} رکورد از تگ تکراری استخراج شد")
                return data
            
            # اگر تگ تکراری نبود، بررسی المان‌هایی که لیست دارند
            for child in root:
                # بررسی کنید که آیا این المان شامل لیستی از زیرالمان‌هاست
                child_dict = self._parse_element(child)
                if child_dict:
                    # بررسی کنید که آیا هر یک از مقادیر، لیست هستند
                    has_list = False
                    for key, value in child_dict.items():
                        if isinstance(value, list) and value and isinstance(value[0], dict):
                            # این یک لیست از دیکشنری‌هاست - باید آن را به رکوردهای جداگانه تبدیل کنیم
                            has_list = True
                            for item in value:
                                if isinstance(item, dict):
                                    # اضافه کردن کلید والد به عنوان پیشوند
                                    new_row = {}
                                    for k, v in item.items():
                                        new_row[f"{key}_{k}"] = v
                                    data.append(new_row)
                    
                    if not has_list:
                        # اگر لیستی نبود، خود المان را اضافه کن
                        data.append(child_dict)
            
            if data:
                logger.info(f"{len(data)} رکورد از ساختار استخراج شد")
                return data
            
            # آخرین راه: کل ریشه را به عنوان یک رکورد
            row = self._parse_element(root)
            if row:
                # بررسی کنید که آیا row دارای لیستی از دیکشنری‌هاست
                final_data = []
                has_list = False
                for key, value in row.items():
                    if isinstance(value, list) and value and isinstance(value[0], dict):
                        has_list = True
                        for item in value:
                            if isinstance(item, dict):
                                new_row = {}
                                for k, v in item.items():
                                    new_row[f"{key}_{k}"] = v
                                final_data.append(new_row)
                
                if has_list:
                    return final_data
                else:
                    return [row]
            
            return []
            
        except Exception as e:
            logger.error(f"خطا در استخراج از XML معمولی: {e}")
            return []
    
    def _parse_element(self, element: ET.Element) -> Dict:
        """پارس یک المان XML به دیکشنری"""
        row = {}
        
        # ویژگی‌ها
        for attr, value in element.attrib.items():
            clean_attr = attr.split('}')[-1] if '}' in attr else attr
            row[f'@{clean_attr}'] = value
        
        # زیرالمان‌ها
        for child in element:
            clean_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            
            if len(child) == 0:  # برگ
                if clean_tag in row:
                    if not isinstance(row[clean_tag], list):
                        row[clean_tag] = [row[clean_tag]]
                    row[clean_tag].append(child.text)
                else:
                    row[clean_tag] = child.text
            else:  # دارای زیرالمان
                child_dict = self._parse_element(child)
                if clean_tag in row:
                    if not isinstance(row[clean_tag], list):
                        row[clean_tag] = [row[clean_tag]]
                    row[clean_tag].append(child_dict)
                else:
                    row[clean_tag] = child_dict
        
        return row
    
    def _flatten_data(self, data: List[Dict]) -> List[Dict]:
        """
        مسطح کردن داده‌ها - هر رکورد در یک سطر جداگانه
        """
        if not data:
            return []
        
        flat_data = []
        
        for item in data:
            flat_row = {}
            self._flatten_dict(item, flat_row)
            flat_data.append(flat_row)
        
        return flat_data
    
    def _flatten_dict(self, d: Dict, result: Dict, prefix: str = ''):
        """مسطح کردن یک دیکشنری"""
        for key, value in d.items():
            # پاک کردن کلید از namespace
            clean_key = key.split('}')[-1] if '}' in key else key
            new_key = f"{prefix}_{clean_key}" if prefix else clean_key
            
            if isinstance(value, dict):
                self._flatten_dict(value, result, new_key)
            elif isinstance(value, list):
                # بررسی اینکه آیا لیست شامل دیکشنری است
                if value and isinstance(value[0], dict):
                    # لیست دیکشنری‌ها - هر کدام را به یک رشته تبدیل کن
                    items = []
                    for sub_item in value:
                        if isinstance(sub_item, dict):
                            parts = []
                            for k, v in sub_item.items():
                                clean_k = k.split('}')[-1] if '}' in k else k
                                if isinstance(v, list):
                                    v = ', '.join(str(x) for x in v if x is not None)
                                parts.append(f"{clean_k}:{v}")
                            items.append('{' + ', '.join(parts) + '}')
                        else:
                            items.append(str(sub_item))
                    result[new_key] = '; '.join(items)
                else:
                    # لیست ساده
                    result[new_key] = ', '.join(str(x) for x in value if x is not None)
            else:
                result[new_key] = value
    
    def get_xml_info(self, xml_file: str) -> Dict:
        """گرفتن اطلاعات ساختار XML"""
        try:
            if self._is_spreadsheet_xml(xml_file):
                return {
                    'root_tag': 'SpreadsheetXML (Excel)',
                    'total_elements': 0,
                    'unique_tags': ['SpreadsheetXML'],
                    'max_depth': 0,
                    'type': 'spreadsheet'
                }
            
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            tags = set()
            max_depth = 0
            
            def traverse(el, depth=0):
                nonlocal max_depth
                tags.add(el.tag)
                max_depth = max(max_depth, depth)
                for child in el:
                    traverse(child, depth + 1)
            
            traverse(root)
            
            return {
                'root_tag': root.tag,
                'total_elements': len(list(root.iter())),
                'unique_tags': list(tags),
                'max_depth': max_depth,
                'type': 'normal'
            }
        except Exception as e:
            logger.error(f"خطا: {e}")
            raise
    
    def convert(self, xml_file: str, output_dir: str = "output", 
                output_name: str = "converted", target_tag: Optional[str] = None,
                output_format: str = "Excel (xlsx)", generate_summary: bool = True) -> Dict[str, str]:
        """
        تبدیل فایل XML به فرمت‌های مختلف
        """
        results = {}
        os.makedirs(output_dir, exist_ok=True)
        
        # استخراج داده
        if self._is_spreadsheet_xml(xml_file):
            raw_data = self._extract_from_spreadsheet_xml(xml_file)
        else:
            raw_data = self._extract_normal_xml(xml_file, target_tag)
        
        if not raw_data:
            raise ValueError("هیچ داده‌ای یافت نشد!")
        
        logger.info(f"📊 {len(raw_data)} رکورد خام استخراج شد")
        
        # مسطح کردن
        self.data = self._flatten_data(raw_data)
        
        if not self.data:
            raise ValueError("خطا در مسطح کردن داده‌ها!")
        
        # استخراج هدرها
        self.headers = list(self.data[0].keys())
        
        logger.info(f"✅ {len(self.data)} رکورد با {len(self.headers)} ستون آماده شد")
        
        # نمایش نمونه
        if self.data:
            logger.info("📝 نمونه داده (اولین رکورد):")
            for key in list(self.data[0].keys())[:6]:
                value = str(self.data[0].get(key, ''))[:50]
                logger.info(f"  {key}: {value}")
        
        # تبدیل به فرمت‌های مختلف
        if output_format in ["Excel (xlsx)", "همه فرمت‌ها"]:
            results["Excel (xlsx)"] = self._save_excel(output_dir, output_name)
        
        if output_format in ["CSV", "همه فرمت‌ها"]:
            results["CSV"] = self._save_csv(output_dir, output_name)
        
        if output_format in ["JSON", "همه فرمت‌ها"]:
            results["JSON"] = self._save_json(output_dir, output_name)
        
        if generate_summary:
            results["گزارش خلاصه"] = self._save_summary(output_dir, output_name)
        
        return results
    
    def _save_excel(self, output_dir: str, output_name: str) -> str:
        """ذخیره به عنوان Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        filepath = os.path.join(output_dir, f"{output_name}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "داده‌های XML"
        
        # هدرها
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=str(header))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # داده‌ها
        for row_idx, row_data in enumerate(self.data, 2):
            for col_idx, header in enumerate(self.headers, 1):
                value = row_data.get(header, '')
                if value is None:
                    value = ''
                elif isinstance(value, (dict, list)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # تنظیم عرض ستون‌ها
        for col in range(1, len(self.headers) + 1):
            max_len = 10
            for row in range(1, min(len(self.data) + 2, 20)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        
        wb.save(filepath)
        logger.info(f"📁 Excel: {filepath}")
        return filepath
    
    def _save_csv(self, output_dir: str, output_name: str) -> str:
        """ذخیره به عنوان CSV"""
        filepath = os.path.join(output_dir, f"{output_name}.csv")
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(self.headers)
            
            for row_data in self.data:
                row = []
                for header in self.headers:
                    value = row_data.get(header, '')
                    if value is None:
                        value = ''
                    elif isinstance(value, (dict, list)):
                        value = str(value)
                    row.append(value)
                writer.writerow(row)
        
        logger.info(f"📁 CSV: {filepath}")
        return filepath
    
    def _save_json(self, output_dir: str, output_name: str) -> str:
        """ذخیره به عنوان JSON"""
        filepath = os.path.join(output_dir, f"{output_name}.json")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"📁 JSON: {filepath}")
        return filepath
    
    def _save_summary(self, output_dir: str, output_name: str) -> str:
        """ذخیره گزارش خلاصه"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        filepath = os.path.join(output_dir, f"{output_name}_summary.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "خلاصه"
        
        summary = [
            ['معیار', 'مقدار'],
            ['تعداد رکوردها', len(self.data)],
            ['تعداد ستون‌ها', len(self.headers)],
            ['نام ستون‌ها', ', '.join(self.headers[:10]) + ('...' if len(self.headers) > 10 else '')],
            ['تاریخ ایجاد', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for row_idx, row in enumerate(summary, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        logger.info(f"📁 Summary: {filepath}")
        return filepath